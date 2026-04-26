# -*- coding: utf-8 -*-

import sys
import importlib
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

import db

import tkinter as tk

import tkinter.font as tkfont

from tkinter import ttk, filedialog, messagebox



try:

    from ctypes import windll

    try:

        windll.shcore.SetProcessDpiAwareness(1)
        windll.shcore.SetProcessDpiAwareness(1)

    except Exception:

        pass

except Exception:

    pass

from parser.pdf_parser import parse_pdf

from config import DB_PATH, SCHEMA

# === INTEGRACIÓN CON FRACCIONADORA ===

FRACC_DB_PATH =  Path(r"C:\Users\osval\Desktop\dev\PDFREADER\GCPDFMK10\GCMK8\fraccionadora.db")
FRACC_APP_PATH = FRACC_DB_PATH.with_suffix(".py")
_PRODUCTOS_REF_CACHE: list[str] | None = None
GCMK8_DIR = FRACC_DB_PATH.parent
for p in (PROJECT_ROOT, GCMK8_DIR):
    if str(p) not in sys.path:
        sys.path.insert(0, str(p))



import re

import unicodedata

from openpyxl import load_workbook

from datetime import datetime, timedelta

import math

import ast

def _table_exists(conn, table_name: str) -> bool:

    return db.table_exists(conn, table_name)



def _norm_nombre_prod(s: str) -> str:

    # normaliza para comparar nombres de producto, ignorando puntuacion del PDF

    base = _norm_txt(s)

    if not base:

        return ""

    base = re.sub(r"[^A-Z0-9]+", " ", base)

    base = re.sub(r"\s+", " ", base).strip()

    return base



def _productos_referencia() -> list[str]:

    global _PRODUCTOS_REF_CACHE

    if _PRODUCTOS_REF_CACHE is not None:

        return _PRODUCTOS_REF_CACHE

    productos: list[str] = []

    path = FRACC_APP_PATH

    if path and path.exists():

        try:

            source = path.read_text(encoding="utf-8")

            tree = ast.parse(source)

            valores = {}

            for node in tree.body:

                if isinstance(node, ast.Assign):

                    for target in node.targets:

                        if isinstance(target, ast.Name) and target.id in ("PRODUCT_ORDER", "MATERIAS_PRIMAS_INICIALES"):

                            valores[target.id] = ast.literal_eval(node.value)

            productos = valores.get("PRODUCT_ORDER") or valores.get("MATERIAS_PRIMAS_INICIALES") or []

        except Exception:

            productos = []

    normalizados = []

    vistos = set()

    for nombre in productos:

        if not nombre:

            continue

        n = str(nombre).strip()

        if n and n not in vistos:

            normalizados.append(n)

            vistos.add(n)

    if not normalizados:

        normalizados = ["Arroz", "Azúcar", "Galleta molida", "Pororó", "Poroto Rojo", "Locro", "Locrillo", "Lenteja"]

    _PRODUCTOS_REF_CACHE = normalizados

    return _PRODUCTOS_REF_CACHE



def _producto_por_desc(desc: str) -> str:

    if not desc:

        return "Otros"

    base = " " + _norm_nombre_prod(desc) + " "

    for nombre in _productos_referencia():

        token = " " + _norm_nombre_prod(nombre) + " "

        if token.strip() and token in base:

            return nombre

    if " ARROZ " in base:

        return "Arroz"

    if (" GALLET" in base or " GALL " in base) and " MOLID" in base:

        return "Galleta molida"

    return "Otros"



def _extraer_gramaje(desc: str) -> int | None:

    """

    Intenta extraer gramaje por línea:

      - '500 g', '500gr', '500 gramos'

      - '1 kg', '1 kilo', '2 kilos'

      - '*400', '* 400', '*400GR'

    """

    if not desc:

        return None

    s = desc.lower()



    m = re.search(r'(\d+)\s*(?:kg|kilo|kilos)\b', s)

    if m:

        return int(m.group(1)) * 1000



    m = re.search(r'(\d+)\s*(?:g|gr|gramo|gramos)\b', s)

    if m:

        return int(m.group(1))



    m = re.search(r'\*\s*(\d{2,4})\s*(?:gr|g)?\b', s)

    if m:

        v = int(m.group(1))

        if 50 <= v <= 5000:

            return v



    return None



def _catalogo_fracc(conn) -> list[tuple[int, str]]:

    """

    Devuelve [(id, name)] del catálogo.

    Usa 'products' si existe; como fallback intenta 'package_stock' si allí hay 'product_id' y 'product_name'.

    """

    cur = conn.cursor()



    # Ruta normal: tabla products

    if _table_exists(conn, "products"):

        cur.execute("SELECT id, name FROM products ORDER BY name;")

        return [(int(r[0]), str(r[1])) for r in cur.fetchall()]



    # Fallback (por si tu base es mínima y no tiene 'products')

    if _table_exists(conn, "package_stock"):

        # Intentar deducir desde package_stock (requiere 'product_id' y 'product_name')

        try:

            cur.execute("SELECT DISTINCT product_id, product_name FROM package_stock WHERE product_id IS NOT NULL AND product_name IS NOT NULL;")

            rows = cur.fetchall()

            if rows:

                return [(int(r[0]), str(r[1])) for r in rows]

        except Exception:

            pass



    # Si no hay forma de obtener catálogo

    return []





def _match_producto(desc: str, productos: list[tuple[int,str]]) -> tuple[int,str] | None:

    """

    Heurística simple: busca el nombre del producto como token dentro de la descripción.

    Si hay varios matches, se queda con el más largo (más específico).

    """

    base = " " + _norm_nombre_prod(desc) + " "

    mejor = None

    for pid, name in productos:

        n = " " + _norm_nombre_prod(name) + " "

        if n in base:

            if (mejor is None) or (len(n) > len(mejor[1])):

                mejor = (pid, name)

    return mejor  # (product_id, product_name) o None


def _buscar_producto_por_nombre(productos: list[tuple[int, str]], nombre_objetivo: str) -> tuple[int, str] | None:
    objetivo = _norm_nombre_prod(nombre_objetivo)
    if not objetivo:
        return None
    for pid, name in productos:
        if _norm_nombre_prod(name) == objetivo:
            return pid, name
    return None



def _stock_paquetes_disponibles(conn, product_id: int, gramaje: int) -> int:

    """

    Lee package_stock(product_id, gramaje)->paquetes. Devuelve 0 si no hay registro.

    """

    cur = conn.cursor()

    if not _table_exists(conn, "package_stock"):

        return 0

    cur.execute("""

        SELECT COALESCE(paquetes, 0)

        FROM package_stock

        WHERE product_id = %s AND gramaje = %s;

    """, (product_id, gramaje))

    row = cur.fetchone()

    return int(row[0]) if row and row[0] is not None else 0



def _mapear_y_chequear_item(desc: str, paquetes_necesarios: int, conn_fracc) -> tuple[str, str, int | None, str | None]:

    """

    Retorna:

      - estado: 'ok' | 'falta' | 'desconocido'

      - detalle: texto breve

      - disp: paquetes disponibles (int) o None si no pudo calcular
      - producto: nombre identificado o None

    """

    if not desc or paquetes_necesarios is None:

        return ("desconocido", "sin datos", None, None)



    gram = _extraer_gramaje(desc)

    if gram is None:

        return ("desconocido", "gramaje no identificado", None, None)



    productos = _catalogo_fracc(conn_fracc)

    match = _match_producto(desc, productos)

    if not match:

        # Fallback: usar nombre canónico cuando la descripción viene abreviada (ej: "GALL. MOLIDA")

        prod_ref = _producto_por_desc(desc)

        if prod_ref and prod_ref != "Otros":

            match = _buscar_producto_por_nombre(productos, prod_ref)

    if not match:

        return ("desconocido", "producto no identificado", None, None)



    pid, pname = match

    disp = _stock_paquetes_disponibles(conn_fracc, pid, gram)

    if disp >= paquetes_necesarios:

        return ("ok", f"{pname} {gram}g: disp {disp} >= req {paquetes_necesarios}", disp, pname)

    else:

        return ("falta", f"{pname} {gram}g: disp {disp} < req {paquetes_necesarios}", disp, pname)


def _column_exists(conn, table, column):

    return column.lower() in {c.lower() for c in db.table_columns(conn, table)}



def migrate_db():

    conn = db.connect("pedidos")

    try:

        if not _column_exists(conn, "orden_compra", "monto_total"):

            conn.execute("ALTER TABLE orden_compra ADD COLUMN monto_total REAL")

            conn.commit()

        if not _column_exists(conn, "orden_compra", "completada"):

            conn.execute("ALTER TABLE orden_compra ADD COLUMN completada INTEGER NOT NULL DEFAULT 0")

            conn.commit()

        if not _column_exists(conn, "orden_item", "enviado"):

            conn.execute("ALTER TABLE orden_item ADD COLUMN enviado INTEGER NOT NULL DEFAULT 0")

            conn.commit()

    finally:

        conn.close()

        



APP_DIR = Path(__file__).resolve().parent

# renombra este archivo si tu plantilla se llama distinto

TEMPLATE_XLSX = APP_DIR / "Copia de NOTA_DE_REMISION_GRANOS_PARA_PEDIDOS(1).xlsx"

OUTPUT_DIR = APP_DIR / "salidas"

from openpyxl import load_workbook

def cargar_cantidades_en_plantilla(template_xlsx: str | Path,

                                   output_xlsx: str | Path,

                                   items: list[dict]):

    """

    Abre la plantilla, localiza columnas de descripción/cantidad (con sinónimos),

    escribe cantidades y guarda en output_xlsx. Recorre todas las hojas hasta encontrar header.

    """

    wb = load_workbook(filename=str(template_xlsx))

    mapa = cantidades_por_descripcion(items)



    ws_target = None

    header_row = col_desc = col_cant = None



    # intenta en todas las hojas; elige la primera que tenga headers válidos

    for ws in wb.worksheets:

        try:

            r, cdesc, ccant = _find_header_indexes(ws)

            ws_target, header_row, col_desc, col_cant = ws, r, cdesc, ccant

            break

        except Exception:

            continue



    if ws_target is None:

        raise RuntimeError("No identifiqué hoja con 'Descripción' y 'Cantidad'.")



    escritos = 0

    for r in range(header_row + 1, ws_target.max_row + 1):

        desc_raw = _cell_value(ws_target, r, col_desc)

        desc_norm = _norm_txt(desc_raw)

        if not desc_norm:

            continue

        if desc_norm in mapa:

            ws_target.cell(row=r, column=col_cant).value = mapa[desc_norm]

            escritos += 1



    Path(output_xlsx).parent.mkdir(parents=True, exist_ok=True)

    wb.save(str(output_xlsx))

    return escritos



# ======= HEADERS ROBUSTOS PARA EXCEL =======

DESC_HEADERS = {

    "DESCRIPCION","DESCRIPCIÓN","DESCRIPCION DEL PRODUCTO","DESCRIPCION PRODUCTO",

    "PRODUCTO","DETALLE","DESCRIPCION ARTICULO","DESCRIPCION ARTÍCULO",

    "ARTICULO","ARTÍCULO","DESCRIPCIÓN DEL PRODUCTO","DESCRIPCION DETALLADA (INCLUIR CANTIDAD O PORCENTAJE DE TOLERANCIA DE QUIEBRA O DE MERMA/ DATOS DE RELEVANCIA DE LA MERCADERIA				"





}

CANT_HEADERS = {

    "CANTIDAD","CANT.","CANT","CANTIDAD SOLICITADA","CANTIDAD PEDIDA","CANTIDAD REQUERIDA",

    "CANT PEDIDA","CANT. PEDIDA","CANT."

}



def _norm_txt(s: str) -> str:

    """Normaliza: sin acentos, sin dobles espacios, mayúsculas, sin puntuación sobrante."""

    if not s:

        return ""

    s = unicodedata.normalize("NFKD", str(s))

    s = "".join(c for c in s if not unicodedata.combining(c))

    s = re.sub(r"\s+", " ", s)

    s = s.strip(" \t\r\n.·,:;")

    return s.upper()



def _cell_value(ws, r, c):

    """Lee valor considerando celdas combinadas."""

    v = ws.cell(row=r, column=c).value

    if v is not None:

        return v

    for rng in ws.merged_cells.ranges:

        if rng.min_row <= r <= rng.max_row and rng.min_col <= c <= rng.max_col:

            return ws.cell(row=rng.min_row, column=rng.min_col).value

    return None



def _find_header_indexes(ws) -> tuple[int, int, int]:

    """

    Devuelve (header_row, col_desc, col_cant) tolerando celdas combinadas y sinónimos.

    Busca en las primeras 50 filas/40 columnas.

    """

    max_r = min(ws.max_row or 50, 50)

    max_c = min(ws.max_column or 40, 40)



    for r in range(1, max_r + 1):

        col_desc = col_cant = None

        for c in range(1, max_c + 1):

            txt = _norm_txt(_cell_value(ws, r, c))

            if not txt:

                continue

            if txt in DESC_HEADERS and col_desc is None:

                col_desc = c

            if txt in CANT_HEADERS and col_cant is None:

                col_cant = c

        if col_desc and col_cant:

            return r, col_desc, col_cant



    raise RuntimeError("No encontré columnas equivalentes a 'Descripción' y 'Cantidad' en la plantilla.")





def _norm_txt(s: str) -> str:

    """Normaliza para comparar: sin acentos, sin dobles espacios, mayúsculas."""

    if not s:

        return ""

    s = unicodedata.normalize("NFKD", s)

    s = "".join(c for c in s if not unicodedata.combining(c))

    s = re.sub(r"\s+", " ", s)

    return s.strip().upper()



def cantidades_por_descripcion(items: list[dict]) -> dict[str, int]:

    """

    items: [{'descripcion': str, 'cantidad': int/float/None}, ...]

    Devuelve un dict normalizado: {DESCRIPCION_NORMALIZADA: cantidad_int}

    """

    q = {}

    for it in items:

        desc = _norm_txt(it.get("descripcion", ""))

        cant = it.get("cantidad")

        if cant is None:

            continue

        try:

            cant = int(round(float(cant)))

        except Exception:

            continue

        q[desc] = cant

    return q



def _gramaje_total_por_paquete(descripcion: str) -> int | None:
    """

    Devuelve el peso total en gramos de UN paquete (caja/fardo) según la descripción.

    Ejemplo: 'AZUCAR EL CACIQUE *1KG (10)' -> 1000g * 10 = 10000g

    """

    if not descripcion:

        return None

    s = descripcion.lower()



    # buscar gramaje

    gramos = None

    m = re.search(r'(\d+)\s*(?:kg|kilo|kilos)\b', s)

    if m:

        gramos = int(m.group(1)) * 1000

    else:

        m = re.search(r'(\d+)\s*(?:g|gr|gramo|gramos)\b', s)

        if m:

            gramos = int(m.group(1))

        else:

            m = re.search(r'\*\s*(\d{2,4})\b', s)

            if m:

                gramos = int(m.group(1))



    if not gramos:

        return None



    # buscar unidades entre paréntesis

    unidades = 1

    m = re.search(r'\((\d+)\)', s)

    if m:

        unidades = int(m.group(1))



    return gramos * unidades


def _categoria_bolsa(desc: str) -> str:
    """
    Clasifica la materia prima base del producto segun descripciA3n.
    """
    prod = _producto_por_desc(desc)
    base = _norm_txt(prod)
    if not base:
        return "OTROS"
    if "ARROZ" in base:
        return "ARROZ"
    if "GALLET" in base and "MOLID" in base:
        return "GALLETAS"
    return "OTROS"


def _peso_bolsa_estandar(desc: str) -> float:
    """
    Determina el peso (kg) de la bolsa base segun el producto.
    """
    tipo = _categoria_bolsa(desc)
    if tipo == "ARROZ":
        return 30.0
    if tipo == "GALLETAS":
        return 25.0
    return 50.0


def _bolsas_necesarias(desc: str, paquetes_requeridos: int, paquetes_disponibles: int | None) -> int | None:
    """
    Calcula cuantas bolsas de materia prima hay que fraccionar para cubrir lo faltante.
    Devuelve None si no se puede estimar (sin paquete disponible o sin gramaje).
    """
    if paquetes_requeridos is None or paquetes_disponibles is None:
        return None

    faltantes = paquetes_requeridos - paquetes_disponibles
    if faltantes <= 0:
        return 0

    gramos_pack = _gramaje_total_por_paquete(desc)
    if not gramos_pack:
        return None

    kg_faltantes = (faltantes * gramos_pack) / 1000.0
    peso_bolsa = _peso_bolsa_estandar(desc)
    if peso_bolsa <= 0:
        return None

    return int(math.ceil(kg_faltantes / peso_bolsa))


def _stock_bolsas_por_producto(conn_fracc) -> dict[str, float] | None:
    """
    Lee raw_stock en fraccionadora y devuelve dict con bolsas equivalentes por producto.
    """
    stock: dict[str, float] = {}
    if conn_fracc is None:
        return None
    try:
        cur = conn_fracc.cursor()
        cur.execute("""
            SELECT p.name, rs.kg
            FROM raw_stock rs
            JOIN products p ON p.id = rs.product_id
        """)
        rows = cur.fetchall()
    except Exception:
        return None
    for name, kg in rows:
        bolsa = _peso_bolsa_estandar(name)
        if bolsa <= 0:
            continue
        prod = _producto_por_desc(name)
        try:
            kg_val = float(kg or 0)
        except Exception:
            kg_val = 0.0
        stock[prod] = stock.get(prod, 0.0) + (kg_val / bolsa)
    return stock


def calcular_peso_total_items(items: list[dict]) -> tuple[float, list[tuple[str, int, float]]]:
    """

    Calcula el peso total en kg del pedido a partir de items ya 'ajustados' (paquetes).

    items: [{'descripcion': str, 'cantidad': int}, ...] donde 'cantidad' son paquetes.

    """

    total_kg = 0.0

    detalle = []

    for it in items:

        desc = it.get("descripcion")

        paquetes = it.get("cantidad") or 0

        gr_total = _gramaje_total_por_paquete(desc)  # gramos por paquete

        if gr_total and paquetes:

            kg_linea = (paquetes * gr_total) / 1000.0

            total_kg += kg_linea

            detalle.append((desc, paquetes, round(kg_linea, 3)))

    return round(total_kg, 3), detalle



def _gramaje_en_gramos(descripcion: str) -> int | None:

    """

    Extrae gramaje en gramos desde la descripción.

    Soporta: '200 g', '200gr', '1 kg', '2 kilos', patrón '*400 (' o '*400GR'.

    Devuelve None si no puede inferir.

    """

    if not descripcion:

        return None

    s = str(descripcion).lower()



    # kg primero

    m = re.search(r'(\d+)\s*(?:kg|kilo|kilos)\b', s)

    if m:

        return int(m.group(1)) * 1000



    # gramos explícitos

    m = re.search(r'(\d+)\s*(?:g|gr|gramo|gramos)\b', s)

    if m:

        return int(m.group(1))



    # patrones tipo "*400 (" o "*400gr"

    m = re.search(r'\*\s*(\d{2,4})\s*(?:gr|g)?\b', s)

    if m:

        v = int(m.group(1))

        if 50 <= v <= 5000:

            return v



    return None





def _unidades_por_paquete(descripcion: str) -> int:

    """

    Regla de negocio:

      - Si el gramaje es <= 300 g => 20 unidades por paquete

      - Si el gramaje es  > 300 g => 10 unidades por paquete

    Si no puede inferir gramaje, asume 10 (conservador).

    """

    gr = _gramaje_en_gramos(descripcion)

    if gr is None:

        return 10

    return 20 if gr <= 300 else 10





def calcular_peso_total_por_oc(nro_oc: str) -> float:

    """

    Calcula el peso total en kg desde la DB para una OC,

    considerando el peso TOTAL de cada paquete (gramaje * unidades del pack).

    """

    conn = db.connect("pedidos")
    try:

        cur = conn.cursor()

        cur.execute("""

            SELECT oi.descripcion AS descripcion, oi.cantidad AS paquetes

            FROM orden_item oi

            JOIN orden_compra oc ON oc.id = oi.oc_id

            WHERE oc.nro_oc = %s AND oi.descripcion IS NOT NULL

        """, (nro_oc,))

        rows = cur.fetchall()

    finally:

        conn.close()



    total_kg = 0.0

    for r in rows:

        paquetes = r["paquetes"] or 0

        gr_total_pack = _gramaje_total_por_paquete(r["descripcion"])  # gramos por paquete (ya incluye (10), (20), etc.)

        if gr_total_pack and paquetes:

            total_kg += (paquetes * gr_total_pack) / 1000.0

    return round(total_kg, 3)



def ajustar_cantidad(descripcion: str, cantidad: float) -> int | None:

    """

    Divide la cantidad según el gramaje del producto y la devuelve como entero.

      - ≤ 300 g  -> cantidad / 20

      -  > 300 g -> cantidad / 10

    Si no detecta peso o cantidad es None/0, devuelve tal cual.

    """

    if cantidad is None or cantidad == 0:

        return cantidad



    desc = (descripcion or "").lower()



    # gramos explícitos: 250g, 250 gr

    m = re.search(r"(\d+)\s*(?:g|gr)\b", desc, flags=re.I)

    if m:

        gramos = int(m.group(1))

        v = cantidad / 20.0 if gramos <= 300 else cantidad / 10.0

        return int(round(v))



    # kilos explícitos: 1kg, 2 kg

    m = re.search(r"(\d+)\s*kg\b", desc, flags=re.I)

    if m:

        gramos = int(m.group(1)) * 1000

        v = cantidad / 20.0 if gramos <= 300 else cantidad / 10.0

        return int(round(v))



    # patrón tipo "*400 (" → infiere gramos

    m = re.search(r"\*\s*(\d{2,4})\s*\(", desc)

    if m:

        valor = int(m.group(1))

        if valor in (200, 250, 300, 400, 500):

            gramos = valor

            v = cantidad / 20.0 if gramos <= 300 else cantidad / 10.0

            return int(round(v))



    return int(round(cantidad))





def ensure_db():

    DB_PATH.parent.mkdir(parents=True, exist_ok=True)

    with open(SCHEMA, "r", encoding="utf-8") as f:

        schema_sql = f.read()

    conn = db.connect("pedidos")

    try:

        db.run_ddl(conn, schema_sql)

    finally:

        conn.close()



def insert_oc(meta, items):

    conn = db.connect("pedidos")

    try:

        cur = conn.cursor()

        monto_total = meta.get("monto_total")

        cur.execute(

            "INSERT INTO orden_compra (nro_oc, sucursal, fecha_pedido, raw_text, monto_total) VALUES (%s,%s,%s,%s,%s) RETURNING id",

            (meta.get("nro_oc"), meta.get("sucursal"), meta.get("fecha_pedido"), meta.get("raw_text"), monto_total)

        )

        cur.execute("SELECT id FROM orden_compra WHERE nro_oc = %s", (meta.get("nro_oc"),))

        row = cur.fetchone()

        if row is None:

            cur.execute(

                "INSERT INTO orden_compra (nro_oc, sucursal, fecha_pedido, raw_text, monto_total) VALUES (%s,%s,%s,%s,%s) RETURNING id",

                (meta.get("nro_oc") or 'SIN_OC', meta.get("sucursal"), meta.get("fecha_pedido"), meta.get("raw_text"), monto_total)

            )

            oc_id = cur.fetchone()[0]

        else:

            oc_id = row["id"]

            cur.execute(
                """
                UPDATE orden_compra
                SET sucursal = %s, fecha_pedido = %s, raw_text = %s, monto_total = %s
                WHERE id = %s
                """,
                (meta.get("sucursal"), meta.get("fecha_pedido"), meta.get("raw_text"), monto_total, oc_id),
            )



        cur.execute("DELETE FROM orden_item WHERE oc_id = %s", (oc_id,))

        for idx, it in enumerate(items, start=1):

            cant_adj = ajustar_cantidad(it.get("descripcion", ""), it.get("cantidad"))

            cur.execute(

                "INSERT INTO orden_item (oc_id, linea, descripcion, cantidad, unidad) VALUES (%s,%s,%s,%s,%s)",

                (oc_id, idx, it.get("descripcion"), cant_adj, it.get("unidad"))

            )

        conn.commit()

        return oc_id

    finally:

        conn.close()

def stock_disponible_por_oc(nro_oc: str) -> int:

    """

    Calcula el total de paquetes disponibles en stock para una OC dada.

    Considera solo los ítems pendientes (no enviados) y usa coincidencia exacta

    de la descripción con la tabla de stock.

    """

    conn = db.connect("pedidos")
    try:

        cur = conn.cursor()

        # Si no existe la tabla de stock, devolvemos 0 para evitar romper la UI.
        if not _table_exists(conn, "stock_productos"):
            return 0

        # Traemos todos los ítems de la OC

        cur.execute("""

            SELECT oi.descripcion, oi.cantidad AS paquetes

            FROM orden_item oi

            JOIN orden_compra oc ON oc.id = oi.oc_id

            WHERE oc.nro_oc = %s

            AND COALESCE(oi.enviado,0) = 0

        """, (nro_oc,))

        items = cur.fetchall()



        total_disp = 0

        for it in items:

            desc = it["descripcion"]

            paquetes = it["paquetes"]



            # Aquí deberías enlazar con tu tabla de stock (ejemplo: stock_productos)
            try:
                cur.execute("""
                    SELECT COALESCE(paquetes_disponibles, 0) AS disp
                    FROM stock_productos
                    WHERE descripcion = %s
                """, (desc,))
                row = cur.fetchone()
                disp = row["disp"] if row else 0
            except Exception:
                # Si la tabla no existe o la estructura cambió, asumimos 0
                disp = 0



            total_disp += min(paquetes, disp)  # lo que realmente se puede cubrir



        return total_disp

    finally:

        conn.close()



class App(tk.Tk):

    def __init__(self):

        super().__init__()

        self.title("Importador de Orden de Compra (PDF → SQLite)")

        self.geometry("980x600")

        self._setup_styles()

        self._configure_high_dpi_scaling()

        self._build_ui()

        ensure_db()

        migrate_db()

        migrate_db()

        self.after(200, self.refresh_resumenes)

        self.after(300, self.refresh_pendientes)





    def _build_ui(self):

        nb = ttk.Notebook(self)

        nb.pack(fill="both", expand=True, padx=12, pady=12)



        # -------- TAB 1: Importar / Parsear --------

        tab_import = ttk.Frame(nb, padding=10)

        nb.add(tab_import, text="1) Importar")

        tab_import.columnconfigure(1, weight=1)

        tab_import.rowconfigure(3, weight=1)



        self.var_pdf = tk.StringVar()

        ttk.Label(tab_import, text="Archivo PDF:").grid(row=0, column=0, sticky="w")

        ttk.Entry(tab_import, textvariable=self.var_pdf, width=90).grid(row=0, column=1, sticky="we", padx=5)

        ttk.Button(tab_import, text="Buscar...", command=self.select_pdf).grid(row=0, column=2)



        bar = ttk.Frame(tab_import, padding=(0, 4))

        bar.grid(row=1, column=0, columnspan=3, sticky="w", pady=8)

        ttk.Button(bar, text="Parsear y Guardar", command=self.parse_and_save).pack(side="left")

        ttk.Button(bar, text="Llenar plantilla XLSX…", command=self.on_fill_template).pack(side="left", padx=8)



        meta_frame = ttk.LabelFrame(tab_import, text="Metadatos")

        meta_frame.grid(row=2, column=0, columnspan=3, sticky="we", pady=5)

        meta_frame.columnconfigure(1, weight=1)

        self.var_oc, self.var_suc, self.var_fecha = tk.StringVar(), tk.StringVar(), tk.StringVar()

        ttk.Label(meta_frame, text="Nro OC:").grid(row=0, column=0, sticky="w")

        ttk.Entry(meta_frame, textvariable=self.var_oc, width=40).grid(row=0, column=1, sticky="we", padx=5)

        ttk.Label(meta_frame, text="Sucursal:").grid(row=1, column=0, sticky="w")

        ttk.Entry(meta_frame, textvariable=self.var_suc, width=40).grid(row=1, column=1, sticky="we", padx=5)

        ttk.Label(meta_frame, text="Fecha:").grid(row=2, column=0, sticky="w")

        ttk.Entry(meta_frame, textvariable=self.var_fecha, width=40).grid(row=2, column=1, sticky="we", padx=5)



        items_frame = ttk.LabelFrame(tab_import, text="Items")

        items_frame.grid(row=3, column=0, columnspan=3, sticky="nsew", pady=5)

        cols = ("linea", "descripcion", "cantidad")

        self.tree = ttk.Treeview(items_frame, columns=cols, show="headings")

        for c, w in (("linea",70), ("descripcion",700), ("cantidad",150)):

            self.tree.heading(c, text=c.capitalize())

            self.tree.column(c, width=w, anchor=("e" if c in ("linea","cantidad") else "w"))

        self.tree.pack(fill="both", expand=True)



        # -------- TAB 2: Resúmenes por OC --------

        # -------- TAB 2: Resúmenes por OC (dividida en dos) --------

        tab_resumen = ttk.Frame(nb, padding=10)

        nb.add(tab_resumen, text="2) Resúmenes por OC")



        # barra superior con acciones

        top_bar = ttk.Frame(tab_resumen, padding=(0, 4))

        top_bar.pack(fill="x", pady=(0,6))

        ttk.Button(top_bar, text="Actualizar", command=self.refresh_resumenes).pack(side="left")

        ttk.Button(top_bar, text="Marcar seleccionado: Pendiente/Completada",

                command=self.toggle_estado_seleccion).pack(side="left", padx=6)
        ttk.Button(top_bar, text="Texto para copiar", command=self._mostrar_texto_pedido)\
            .pack(side="left", padx=6)
        ttk.Button(top_bar, text="Eliminar importación", command=self._eliminar_oc_seleccionada)\
            .pack(side="right", padx=6)



        # panel dividido

        paned = ttk.Panedwindow(tab_resumen, orient="horizontal")

        paned.pack(fill="both", expand=True)



        # ----- Izquierda: lista/resumen de OCs

        left = ttk.Frame(paned)

        paned.add(left, weight=1)



        cols2 = ("nro_oc","sucursal","fecha","items","paquetes","monto","estado")

        self.tree_oc = ttk.Treeview(left, columns=cols2, show="headings", height=12)

        for c, w in (

            ("nro_oc",140),

            ("sucursal",160),

            ("fecha",120),

            ("items",80),

            ("paquetes",100),

            ("monto",120),

            ("estado",120)

        ):

            heading = "MONTO (Gs)" if c == "monto" else c.upper()

            anchor = "center" if c in ("items","paquetes","disponibles","estado") else "w"

            if c == "monto":

                anchor = "e"

            self.tree_oc.heading(c, text=heading)

            self.tree_oc.column(c, width=w, anchor=anchor)

        self.tree_oc.pack(fill="both", expand=True)



        # vincular selección para actualizar el detalle de la derecha

        self.tree_oc.bind("<<TreeviewSelect>>", self._on_select_oc)



        # ----- Derecha: detalle de paquetes de la OC seleccionada

        right = ttk.Frame(paned)

        paned.add(right, weight=1)



        info_box = ttk.LabelFrame(right, text="Información de la OC")

        info_box.pack(fill="x", padx=4, pady=(0,6))

        self.lbl_info = ttk.Label(info_box, text="NRO: -   |   FECHA: -   |   ITEMS: -   |   SUC: -")

        self.lbl_info.pack(anchor="w", padx=8, pady=4)

        # Badge de dias transcurridos desde la fecha de pedido
        self.lbl_dias = tk.Label(
            info_box,
            text="DIAS: -",
            anchor="w",
            padx=8,
            pady=6,
            font=("Segoe UI", 10, "bold"),
            bg="#e5e7eb",
            fg="#111111",
        )
        self.lbl_dias.pack(fill="x", padx=8, pady=(0, 6))



        cols_det = ("linea", "descripcion", "paquetes", "enviado")

        self.tree_oc_detalle = ttk.Treeview(right, columns=cols_det, show="headings", height=12)

        self.tree_oc_detalle.heading("linea", text="LINEA")

        self.tree_oc_detalle.heading("descripcion", text="DESCRIPCIÓN")

        self.tree_oc_detalle.heading("paquetes", text="PAQUETES")

        self.tree_oc_detalle.heading("enviado", text="ENVIADO")

        self.tree_oc_detalle.column("linea", width=70, anchor="center")

        self.tree_oc_detalle.column("descripcion", width=440, anchor="w")

        self.tree_oc_detalle.column("paquetes", width=110, anchor="e")

        self.tree_oc_detalle.column("enviado", width=100, anchor="center")

        self.tree_oc_detalle.pack(fill="both", expand=True)

        self.tree_oc_detalle.bind("<Button-1>", self._on_tree_detalle_click)

        # colores para disponibilidad por ítem

        self.tree_oc_detalle.tag_configure("ok",          foreground="#0a7d28")  # verde

        self.tree_oc_detalle.tag_configure("falta",       foreground="#c1121f")  # rojo

        self.tree_oc_detalle.tag_configure("desconocido", foreground="#a17c00")  # ámbar

        self.tree_oc_detalle.tag_configure("sent",        foreground="#111111")  # enviado/completada





        # -------- TAB 3: Pendientes acumulados --------

# -------- TAB 3: Pendientes acumulados --------

        tab_pend = ttk.Frame(nb, padding=10)

        nb.add(tab_pend, text="3) Pendientes acumulados")

        # -------- TAB 4: Plan producción --------
        tab_plan = ttk.Frame(nb, padding=10)
        nb.add(tab_plan, text="4) Plan producción")



        top_bar2 = ttk.Frame(tab_pend, padding=(0, 4))

        top_bar2.pack(fill="x", pady=(0,6))

        ttk.Button(top_bar2, text="Actualizar", command=self.refresh_pendientes).pack(side="left")
        self.var_total_pend_gs = tk.StringVar(value="0")
        ttk.Label(top_bar2, text="Total pendientes (Gs):").pack(side="right")
        ttk.Label(top_bar2, textvariable=self.var_total_pend_gs, font=("Segoe UI", 10, "bold")).pack(side="right", padx=(0,10))

        self.frm_pend_sucursales = ttk.Frame(tab_pend, padding=(0, 0, 0, 8))
        self.frm_pend_sucursales.pack(fill="x")
        self._pend_sucursal_vars = {}
        self.var_pend_total_general = tk.StringVar(value="Sin datos globales.")
        self.lbl_pend_total_general = ttk.Label(
            tab_pend,
            textvariable=self.var_pend_total_general,
            font=("Segoe UI", 10, "bold"),
            justify="left",
        )
        self.lbl_pend_total_general.pack(fill="x", pady=(0, 8))



        # columnas: descripción (con diagnóstico), necesario, disponible

        cols3 = ("descripcion","necesario","disponible","bolsas")
        self.tree_pend = ttk.Treeview(tab_pend, columns=cols3, show="headings", height=12)

        self.tree_pend.heading("descripcion", text="DESCRIPCIÓN")

        self.tree_pend.heading("necesario", text="NECESARIO")

        self.tree_pend.heading("disponible", text="DISPONIBLE")

        self.tree_pend.heading("bolsas", text="BOLSAS NEC.")
        self.tree_pend.column("descripcion", width=640, anchor="w")
        self.tree_pend.column("necesario", width=110, anchor="e")

        self.tree_pend.column("disponible", width=110, anchor="e")

        self.tree_pend.column("bolsas", width=130, anchor="e")
        self.tree_pend.pack(fill="both", expand=True)

        # resumen de bolsas necesarias por materia prima
        productos_ref = _productos_referencia()
        resumen_items = list(dict.fromkeys(productos_ref + ["Otros"]))
        self.var_bolsas = {nombre: tk.StringVar(value="0") for nombre in resumen_items}
        self.lbl_bolsas_estado = {}

        cards = ttk.Frame(tab_pend, padding=(0, 8))
        cards.pack(fill="x")
        cols = min(3, max(1, len(resumen_items)))
        for c in range(cols):
            cards.columnconfigure(c, weight=1, uniform="bolsas")

        for idx, nombre in enumerate(resumen_items):
            fila, col = divmod(idx, cols)
            peso = _peso_bolsa_estandar(nombre)
            titulo = f"{nombre} (bolsa {int(peso)} kg)" if peso else nombre
            box = ttk.LabelFrame(cards, text=titulo, padding=10)
            box.grid(row=fila, column=col, sticky="nsew", padx=5, pady=5)
            ttk.Label(box, textvariable=self.var_bolsas[nombre], font=("Segoe UI", 18, "bold")).pack(anchor="center")
            style_obj = getattr(self, "_style", None) or ttk.Style(self)
            bg_card = style_obj.lookup("TLabelframe", "background") or "#ffffff"
            estado_lbl = tk.Label(box, text="SIN DATOS", font=("Segoe UI", 10, "bold"), fg="#a17c00", bg=bg_card)
            estado_lbl.pack(anchor="center", pady=(4,0))
            self.lbl_bolsas_estado[nombre] = estado_lbl



        # colores (mismos que en la pestaña de OC)

        self.tree_pend.tag_configure("ok",     foreground="#0a7d28")  # verde

        self.tree_pend.tag_configure("falta",  foreground="#c1121f")  # rojo

        self.tree_pend.tag_configure("desconocido", foreground="#a17c00")  # ámbar

        # etiquetas destacadas para estado de OC
        pill_green_bg = "#b9e6b9"
        pill_green_fg = "#0b4f2c"
        pill_red_bg = "#d63a3a"
        pill_red_fg = "#ffffff"

        # Nota: el fondo lo maneja el zebra; aquí solo color de texto y peso
        self.tree_oc.tag_configure(
            "estado_completada",
            foreground=pill_green_fg,
            font=("Segoe UI", 10, "bold"),
        )

        self.tree_oc.tag_configure(
            "estado_pendiente",
            foreground=pill_red_fg,
            font=("Segoe UI", 10, "bold"),
        )

        # zebra por estado
        comp_even = "#b2e6b2"   # verde más vivo
        comp_odd  = "#90d690"   # verde alterno
        pend_even = "#e05a5a"   # rojo más vivo
        pend_odd  = "#c63f3f"   # rojo alterno

        self.tree_oc.tag_configure(
            "estado_completada_even",
            background=comp_even,
            foreground=pill_green_fg,
            font=("Segoe UI", 10, "bold"),
        )

        self.tree_oc.tag_configure(
            "estado_completada_odd",
            background=comp_odd,
            foreground=pill_green_fg,
            font=("Segoe UI", 10, "bold"),
        )

        self.tree_oc.tag_configure(
            "estado_pendiente_even",
            background=pend_even,
            foreground=pill_red_fg,
            font=("Segoe UI", 10, "bold"),
        )

        self.tree_oc.tag_configure(
            "estado_pendiente_odd",
            background=pend_odd,
            foreground=pill_red_fg,
            font=("Segoe UI", 10, "bold"),
        )

        for tv in (self.tree, self.tree_oc, self.tree_oc_detalle, self.tree_pend):

            self._apply_treeview_striping(tv)

        # Construir plan de producción
        self._build_tab_plan(tab_plan)



    def _build_tab_plan(self, tab):
        top = ttk.Frame(tab, padding=(0, 4))
        top.pack(fill="x", pady=(0, 6))
        ttk.Label(top, text="Ventana días consumo:").pack(side="left")
        self.ent_plan_window = ttk.Entry(top, width=8)
        self.ent_plan_window.insert(0, "30")
        self.ent_plan_window.pack(side="left", padx=4)
        ttk.Button(top, text="Refrescar plan", command=self._refresh_planificador).pack(side="left", padx=6)
        ttk.Label(top, text="Consumo en días con movimiento; incluye pendientes de OC.").pack(side="left", padx=6)

        cols = ("producto","stock","cons_dia","dias_rest","oc_kg","dias_con_oc","dias_act")
        self.tree_plan = ttk.Treeview(tab, columns=cols, show="headings", height=12)
        headers = {
            "producto": "Producto",
            "stock": "Stock kg",
            "cons_dia": "Cons/día (kg)",
            "dias_rest": "Días restantes",
            "oc_kg": "Kg en OC",
            "dias_con_oc": "Días con OC",
            "dias_act": "Días con consumo",
        }
        widths = {"producto":250,"stock":110,"cons_dia":130,"dias_rest":120,"oc_kg":110,"dias_con_oc":130,"dias_act":120}
        for c in cols:
            self.tree_plan.heading(c, text=headers[c])
            anchor = "w" if c=="producto" else ("e" if c in ("stock","cons_dia","oc_kg") else "center")
            self.tree_plan.column(c, width=widths[c], anchor=anchor)
        self.tree_plan.tag_configure("crit", foreground="#b91c1c")
        self.tree_plan.tag_configure("warn", foreground="#d97706")
        self.tree_plan.tag_configure("ok", foreground="#0f766e")
        self.tree_plan.pack(fill="both", expand=True, pady=(0,6))
        self.lbl_plan_info = ttk.Label(tab, text="Cargando plan...", foreground="#374151")
        self.lbl_plan_info.pack(anchor="w", pady=(0,4))

        cal_frame = ttk.LabelFrame(tab, text="Calendario (días hábiles; primero Azúcar/Arroz, luego resto en el mismo día)")
        cal_frame.pack(fill="both", expand=True, pady=(6,0))
        top_cal = ttk.Frame(cal_frame, padding=(0,4))
        top_cal.pack(fill="x")
        ttk.Label(top_cal, text="Días para entregar OC:").pack(side="left")
        self.ent_plan_days = ttk.Entry(top_cal, width=6)
        self.ent_plan_days.insert(0, "5")
        self.ent_plan_days.pack(side="left", padx=4)
        ttk.Button(top_cal, text="Recalcular calendario", command=self._refresh_planificador).pack(side="left", padx=6)

        week_cols = ("dia","fecha","producto","kg_oc","bolsas_eq")
        self.tree_week = ttk.Treeview(cal_frame, columns=week_cols, show="headings", height=10)
        for c, txt, w, anch in [
            ("dia","Día",80,"center"),
            ("fecha","Fecha",110,"center"),
            ("producto","Producto",220,"w"),
            ("kg_oc","Kg OC",110,"e"),
            ("bolsas_eq","Bolsas (eq)",110,"e"),
        ]:
            self.tree_week.heading(c, text=txt)
            self.tree_week.column(c, width=w, anchor=anch)
        yscroll = ttk.Scrollbar(cal_frame, orient="vertical", command=self.tree_week.yview)
        self.tree_week.configure(yscrollcommand=yscroll.set)
        self.tree_week.pack(side="left", fill="both", expand=True, padx=(6,0), pady=6)
        yscroll.pack(side="right", fill="y", padx=(0,6), pady=6)
        self.lbl_week_info = ttk.Label(cal_frame, text="Plan por gramaje: Azúcar/Arroz primero (un gramaje por día), luego el resto agrupado en el siguiente día hábil.")
        self.lbl_week_info.pack(anchor="w", padx=6, pady=(0,6))
        self._refresh_planificador()

    def _refresh_planificador(self):
        try:
            from GCMK8 import planificador_produccion as plan_mod
            importlib.reload(plan_mod)
        except Exception as exc:
            messagebox.showerror("Plan producción", f"No se pudo cargar planificador_produccion.py: {exc}")
            return
        try:
            ventana = int((self.ent_plan_window.get() or "30").strip())
            if ventana <= 0:
                ventana = 30
        except Exception:
            ventana = 30
            self.ent_plan_window.delete(0, tk.END)
            self.ent_plan_window.insert(0, "30")

        try:
            rows = plan_mod.cargar_plan(ventana)
            prod_map = plan_mod._map_products()
            oc_lines = plan_mod.pendientes_por_gramaje(prod_map)
        except Exception as exc:
            messagebox.showerror("Plan producción", str(exc))
            return

        for iid in self.tree_plan.get_children():
            self.tree_plan.delete(iid)

        crit = warn = 0
        cap_map = {}
        for r in rows:
            dias_base = r.get("dias_restantes")
            dias_oc = r.get("dias_con_oc")
            tag = "ok"
            try:
                val = float(dias_oc if dias_oc is not None else dias_base if dias_base is not None else 1e9)
                if val <= 7:
                    tag = "crit"; crit += 1
                elif val <= 15:
                    tag = "warn"; warn += 1
            except Exception:
                tag = "ok"
            cap_map[r.get("producto","")] = float(r.get("prod_dia_est", 0) if isinstance(r, dict) else r.get("consumo_diario",0))
            self.tree_plan.insert("", "end", values=(
                r.get("producto","-"),
                f"{float(r.get('stock_kg') or 0):.2f}",
                f"{float(r.get('consumo_diario') or 0):.2f}",
                "-" if dias_base is None else f"{float(dias_base):.1f}",
                f"{float(r.get('oc_kg') or 0):.2f}",
                "-" if dias_oc is None else f"{float(dias_oc):.1f}",
                str(r.get("dias_activos") or 0),
            ), tags=(tag,))
        self._apply_treeview_striping(self.tree_plan)
        self.lbl_plan_info.config(text=f"{len(rows)} productos | Críticos (<=7d): {crit} | Aviso (<=15d): {warn}")

        # Calendario
        for iid in self.tree_week.get_children():
            self.tree_week.delete(iid)
        try:
            cal_days = int((self.ent_plan_days.get() or "5").strip())
            if cal_days <= 0:
                cal_days = 5
        except Exception:
            cal_days = 5
            self.ent_plan_days.delete(0, tk.END)
            self.ent_plan_days.insert(0, "5")

        dias_es = {0:"Lun",1:"Mar",2:"Mié",3:"Jue",4:"Vie",5:"Sáb",6:"Dom"}
        today = datetime.now().date()
        days_list = []
        offset = 0
        while len(days_list) < cal_days:
            d = today + timedelta(days=offset)
            if d.weekday() < 5:
                days_list.append(d)
            offset += 1

        def _prio(line):
            name = (line.get("producto") or "").lower().replace("á","a").replace("í","i").replace("ó","o").replace("ú","u")
            if "azucar" in name:
                return 0
            if "arroz" in name:
                return 1
            return 2

        prio_lines = []
        other_lines = []
        for ln in sorted(oc_lines, key=_prio):
            n = (ln.get("producto") or "").lower()
            if "azucar" in n or "arroz" in n:
                prio_lines.append(ln)
            else:
                other_lines.append(ln)

        week_rows = []
        day_idx = 0
        # Asignar prioritarios: un gramaje por día
        for ln in prio_lines:
            prod = ln.get("producto","-")
            gram = ln.get("gramaje")
            rem = float(ln.get("kg") or 0.0)
            while rem > 1e-9 and day_idx < len(days_list):
                d = days_list[day_idx]
                cap = float(cap_map.get(prod, 0.0) or rem)
                if cap <= 0:
                    cap = rem
                kg_today = min(cap, rem)
                bolsa = _peso_bolsa_estandar(prod) if callable(globals().get("_peso_bolsa_estandar")) else 0
                bolsas_eq = "-" if not bolsa else f"{kg_today / bolsa:.2f}"
                semana = dias_es.get(d.weekday(), d.strftime("%a"))
                label = prod if gram is None else f"{prod} {int(gram)}g"
                week_rows.append((semana, d.isoformat(), label, f"{kg_today:.2f}", bolsas_eq))
                rem -= kg_today
                day_idx += 1
            if day_idx >= len(days_list):
                break

        # Luego, agrupar el resto en el siguiente día disponible
        if day_idx < len(days_list):
            d = days_list[day_idx]
            semana = dias_es.get(d.weekday(), d.strftime("%a"))
            for ln in other_lines:
                prod = ln.get("producto","-")
                gram = ln.get("gramaje")
                rem = float(ln.get("kg") or 0.0)
                if rem <= 1e-9:
                    continue
                bolsa = _peso_bolsa_estandar(prod) if callable(globals().get("_peso_bolsa_estandar")) else 0
                bolsas_eq = "-" if not bolsa else f"{rem / bolsa:.2f}"
                label = prod if gram is None else f"{prod} {int(gram)}g"
                week_rows.append((semana, d.isoformat(), label, f"{rem:.2f}", bolsas_eq))

        for w in week_rows:
            self.tree_week.insert("", "end", values=w)
        if not week_rows:
            self.lbl_week_info.config(text="No hay pendientes de OC para planificar. Importa OCs o revisa gramajes/unidades.")
        else:
            self.lbl_week_info.config(text="Azúcar/Arroz primero (un gramaje por día). El resto se agrupa en el siguiente día hábil.")

    def _setup_styles(self):

        style = ttk.Style(self)

        if "clam" in style.theme_names():

            style.theme_use("clam")



        base_bg = "#f7fbf5"

        card_bg = "#ffffff"

        primary_text = "#12326b"

        accent_blue = "#0d4ba0"

        zebra_alt = "#f0f4ff"



        self.configure(bg=base_bg)

        for fname in ("TkDefaultFont", "TkTextFont", "TkHeadingFont", "TkMenuFont"):

            try:

                tkfont.nametofont(fname).configure(family="Segoe UI", size=10)

            except tk.TclError:

                pass



        style.configure("TFrame", background=base_bg)

        style.configure("TLabel", foreground=primary_text, background=base_bg)

        style.configure("TNotebook", background=base_bg, padding=6)

        style.configure(

            "TNotebook.Tab",

            padding=(18, 8),

            foreground=primary_text,

            background=card_bg,

        )

        style.map(

            "TNotebook.Tab",

            background=[("selected", card_bg), ("active", "#eef3ff")],

            foreground=[("selected", accent_blue)],

        )

        style.configure("TButton", padding=(14, 6), background=accent_blue, foreground="#ffffff")

        style.map("TButton", background=[("active", "#125ec9"), ("pressed", "#0b3f79")])

        style.configure("TLabelframe", background=card_bg, padding=10, borderwidth=1, relief="solid")

        style.configure("TLabelframe.Label", background=card_bg, foreground=primary_text)

        style.configure(

            "Treeview",

            background=card_bg,

            fieldbackground=card_bg,

            borderwidth=0,

            rowheight=24,

            font=("Segoe UI", 10),

        )

        style.configure(

            "Treeview.Heading",

            font=("Segoe UI", 10, "bold"),

            padding=6,

            background=card_bg,

            foreground=accent_blue,

        )

        style.map(

            "Treeview",

            background=[("selected", "#d9e5ff")],

            foreground=[("selected", primary_text)],

        )



        self._style = style

        self._stripe_colors = (card_bg, zebra_alt)



    def _apply_treeview_striping(self, tree):

        if not tree:
            return

        colors = getattr(self, "_stripe_colors", ("#ffffff", "#f6f8fc"))
        tree.tag_configure("evenrow", background=colors[0])
        tree.tag_configure("oddrow", background=colors[1])

        oc_tree = getattr(self, "tree_oc", None)
        estado_zebra_tags = {
            "estado_completada": ("estado_completada_even", "estado_completada_odd"),
            "estado_pendiente": ("estado_pendiente_even", "estado_pendiente_odd"),
        }

        for idx, iid in enumerate(tree.get_children()):
            base_tags = [t for t in tree.item(iid, "tags") if t not in (
                "evenrow",
                "oddrow",
                "estado_completada_even",
                "estado_completada_odd",
                "estado_pendiente_even",
                "estado_pendiente_odd",
            )]

            if tree is oc_tree:
                zebra_tag = None
                for estado_tag, zebra_opts in estado_zebra_tags.items():
                    if estado_tag in base_tags:
                        zebra_tag = zebra_opts[0] if idx % 2 == 0 else zebra_opts[1]
                        break
                if zebra_tag is None:
                    zebra_tag = "evenrow" if idx % 2 == 0 else "oddrow"
                base_tags = [zebra_tag] + base_tags
            else:
                base_tags.append("evenrow" if idx % 2 == 0 else "oddrow")

            tree.item(iid, tags=tuple(base_tags))

    def _dias_desde_fecha(self, fecha_str: str) -> int | None:
        """
        Devuelve dias transcurridos desde fecha_str (formato YYYY-MM-DD o DD/MM/YYYY).
        """
        if not fecha_str:
            return None
        fecha_str = str(fecha_str).strip()
        formatos = ("%Y-%m-%d", "%d/%m/%Y")
        for fmt in formatos:
            try:
                dt = datetime.strptime(fecha_str[:10], fmt)
                return (datetime.now().date() - dt.date()).days
            except Exception:
                continue
        return None

    def _set_dias_badge(self, dias: int | None, completada: bool = False):
        """
        Actualiza el badge visual de dias en la tarjeta de info de la OC.
        Si la OC esta completada, muestra un mensaje de entregado en verde.
        """
        if not hasattr(self, "lbl_dias"):
            return

        if completada:
            bg, fg, txt = "#b9e6b9", "#0b4f2c", "ENTREGADA"
        else:
            colors = {
                1: ("#d7ecff", "#0a3d62"),  # dia 0
                2: ("#ffe8b3", "#7a4a00"),  # dia 1
                3: ("#ffc285", "#8a3b00"),  # dia 2
                4: ("#ffb3b3", "#8b0000"),  # 3+ dias
            }

            if dias is None or dias < 0:
                bg, fg = "#e5e7eb", "#111111"
                txt = "DIAS: -"
            else:
                nivel = min(4, dias + 1)
                bg, fg = colors.get(nivel, ("#e5e7eb", "#111111"))
                fuego = " 🔥" if nivel == 4 else ""
                txt = f"DIAS: {dias}{fuego}"

        try:
            self.lbl_dias.config(text=txt, bg=bg, fg=fg)
        except Exception:
            pass

    def _fmt_monto(self, value):

        """Formatea montos con miles usando puntos para mayor legibilidad."""

        try:

            return f"{float(value):,.0f}".replace(",", ".")

        except Exception:

            return ""


    def _set_item_enviado(self, item_id: int, flag: int):

        conn = db.connect("pedidos")

        try:

            cur = conn.cursor()

            cur.execute("UPDATE orden_item SET enviado = %s WHERE id = %s", (1 if flag else 0, item_id))

            conn.commit()

        finally:

            conn.close()



    def _on_tree_detalle_click(self, event):

        region = self.tree_oc_detalle.identify("region", event.x, event.y)

        if region != "cell":

            return

        column = self.tree_oc_detalle.identify_column(event.x)

        if column != "#4":

            return

        row_id = self.tree_oc_detalle.identify_row(event.y)

        if not row_id:

            return "break"

        nro = self._get_selected_oc()

        if not nro:

            return "break"

        current = self.tree_oc_detalle.set(row_id, "enviado")

        new_flag = 0 if current == "[x]" else 1

        try:

            item_id = int(row_id)

        except ValueError:

            return "break"

        self._set_item_enviado(item_id, new_flag)

        self._refresh_oc_info_and_detail(nro)

        self.refresh_pendientes()

        return "break"



    def _configure_high_dpi_scaling(self):

        """

        Ajusta el factor de escala de Tk en pantallas HiDPI para evitar que la UI se vea borrosa.

        """

        try:

            self.update_idletasks()

            px_per_inch = self.winfo_fpixels("1i")

            scaling = max(1.0, min(1.7, px_per_inch / 72.0))

            self.tk.call("tk", "scaling", scaling)

        except tk.TclError:

            pass



    def on_fill_template(self):

                if not getattr(self, "items_full", None):

                    messagebox.showwarning("Sin datos", "Primero carga un PDF.")

                    return

                tpl = filedialog.askopenfilename(title="Selecciona plantilla XLSX",

                                                filetypes=[("Excel", "*.xlsx")])

                if not tpl:

                    return

                out = filedialog.asksaveasfilename(title="Guardar como",

                                                defaultextension=".xlsx",

                                                filetypes=[("Excel", "*.xlsx")])

                if not out:

                    return

                # items_full tiene {'linea','desc','qty'}; conviértelo a la forma que espera la función

                items = [{"descripcion": r["desc"], "cantidad": r["qty"]} for r in self.items_full]

                try:

                    escritos = cargar_cantidades_en_plantilla(tpl, out, items)

                except Exception as e:

                    messagebox.showerror("Error", str(e)); return

                messagebox.showinfo("Listo", f"Actualizadas {escritos} filas en {out}")



    def _on_select_oc(self, _evt=None):

        nro = self._get_selected_oc()

        if not nro:

            return

        self._refresh_oc_info_and_detail(nro)



    def _refresh_oc_info_and_detail(self, nro_oc: str):

        # Franja superior

        conn = db.connect("pedidos")
        try:

            cur = conn.cursor()

            cur.execute("""

                SELECT oc.nro_oc, oc.sucursal, oc.fecha_pedido, oc.monto_total, oc.completada,

                    COUNT(oi.id) AS items

                FROM orden_compra oc

                LEFT JOIN orden_item oi ON oi.oc_id = oc.id

                WHERE oc.nro_oc = %s

                GROUP BY oc.id

            """, (nro_oc,))

            row = cur.fetchone()

        finally:

            conn.close()



        completada = bool(row["completada"]) if row else False

        dias_val = None

        if row:

            monto_txt = ""

            if row["monto_total"] is not None:

                monto_txt = f"   |   MONTO: {self._fmt_monto(row['monto_total'])} Gs"

            self.lbl_info.config(

                text=f"NRO: {row['nro_oc'] or '-'}   |   FECHA: {row['fecha_pedido'] or '-'}   |   ITEMS: {row['items']}   |   SUC: {row['sucursal'] or '-'}{monto_txt}"

            )

            dias_val = self._dias_desde_fecha(row["fecha_pedido"])

        else:

            self.lbl_info.config(text="NRO: -   |   FECHA: -   |   ITEMS: -   |   SUC: -")

        self._set_dias_badge(dias_val, completada)



        try:

            peso_kg = calcular_peso_total_por_oc(nro_oc)

            self.lbl_info.config(text=self.lbl_info.cget("text") + f"   |   PESO: {peso_kg} kg")

        except Exception:

            pass



        # Detalle

        self.tree_oc_detalle.delete(*self.tree_oc_detalle.get_children())



        conn = db.connect("pedidos")
        try:

            cur = conn.cursor()

            cur.execute("""

                SELECT

                    oi.id,

                    oi.linea,

                    TRIM(CAST(oi.descripcion AS TEXT)) AS descripcion,

                    COALESCE(oi.cantidad, 0) AS paquetes,

                    COALESCE(oi.enviado, 0) AS enviado

                FROM orden_item oi

                JOIN orden_compra oc ON oc.id = oi.oc_id

                WHERE oc.nro_oc = %s AND oi.descripcion IS NOT NULL

                ORDER BY oi.linea ASC, descripcion ASC

            """, (nro_oc,))

            rows = cur.fetchall()

        finally:

            conn.close()



        fr_conn = None

        try:

            fr_conn = db.connect("fraccionadora")

        except Exception:

            fr_conn = None



        for r in rows:

            item_id = r["id"]

            desc = r["descripcion"]

            req = int(round(r["paquetes"] or 0))

            linea_txt = "" if r["linea"] is None else str(r["linea"])

            enviado = bool(r["enviado"])



            if completada or enviado:

                texto_desc = desc

                tag = ("sent",)

            else:

                estado, info, _disp, _ = _mapear_y_chequear_item(desc, req, fr_conn) if fr_conn else ("desconocido", "", None, None)

                texto_desc = desc if not info else f"{desc}  -  [{info}]"

                tag = (estado,)



            check_txt = "[x]" if enviado else "[ ]"

            self.tree_oc_detalle.insert(

                "",

                "end",

                iid=str(item_id),

                values=(linea_txt, texto_desc, req, check_txt),

                tags=tag,

            )



        if fr_conn:

            fr_conn.close()

        self._apply_treeview_striping(self.tree_oc_detalle)



    def refresh_resumenes(self):

        # recordar selección previa

        sel = self.tree_oc.selection()

        sel_nro = None

        if sel:

            vals = self.tree_oc.item(sel[0], "values")

            sel_nro = vals[0] if vals else None



        conn = db.connect("pedidos")
        try:

            cur = conn.cursor()

            cur.execute("""

                SELECT oc.nro_oc, oc.sucursal, oc.fecha_pedido, oc.monto_total,

                    COUNT(oi.id) AS items,

                    COALESCE(SUM(oi.cantidad),0) AS paquetes,

                    CASE WHEN oc.completada=1 THEN 'COMPLETADA' ELSE 'PENDIENTE' END AS estado

                FROM orden_compra oc

                LEFT JOIN orden_item oi ON oi.oc_id = oc.id

                GROUP BY oc.id

                ORDER BY oc.fecha_pedido DESC, oc.nro_oc

            """)

            rows = cur.fetchall()

        finally:

            conn.close()



        self.tree_oc.delete(*self.tree_oc.get_children())

        to_select_iid = None



        for r in rows:

            monto_txt = "" if r["monto_total"] is None else self._fmt_monto(r["monto_total"])

            estado_display = "  COMPLETADA  " if r["estado"] == "COMPLETADA" else "  PENDIENTE  "

            estado_tag = "estado_completada" if r["estado"] == "COMPLETADA" else "estado_pendiente"

            iid = self.tree_oc.insert("", "end", values=(

                r["nro_oc"] or "SIN_OC",

                r["sucursal"] or "",

                r["fecha_pedido"] or "",

                r["items"], r["paquetes"], monto_txt,

                estado_display

            ), tags=(estado_tag,))


            if sel_nro and r["nro_oc"] == sel_nro:

                to_select_iid = iid



        self._apply_treeview_striping(self.tree_oc)



        if to_select_iid:

            self.tree_oc.selection_set(to_select_iid)

            self.tree_oc.see(to_select_iid)

            self._refresh_oc_info_and_detail(sel_nro)

        else:

            self.lbl_info.config(text="NRO: -   |   FECHA: -   |   ITEMS: -   |   SUC: -")

            self._set_dias_badge(None, completada=False)

            self.tree_oc_detalle.delete(*self.tree_oc_detalle.get_children())



    def _get_selected_oc(self):

        sel = self.tree_oc.selection()

        if not sel:

            return None

        vals = self.tree_oc.item(sel[0], "values")

        return vals[0]  # nro_oc



    def toggle_estado_seleccion(self):

        nro = self._get_selected_oc()

        if not nro:

            messagebox.showwarning("Selecciona una fila", "Elegí una OC para alternar su estado.")

            return

        conn = db.connect("pedidos")

        try:

            cur = conn.cursor()

            cur.execute("SELECT completada FROM orden_compra WHERE nro_oc = %s", (nro,))

            row = cur.fetchone()

            if row is None:

                messagebox.showerror("No existe", f"No encontré la OC {nro}")

                return

            nuevo = 0 if row[0] == 1 else 1

            cur.execute("UPDATE orden_compra SET completada = %s WHERE nro_oc = %s", (nuevo, nro))

            conn.commit()

        finally:

            conn.close()

        self.refresh_resumenes()

        self.refresh_pendientes()  # para que se actualice también el agregado

    def _eliminar_oc_seleccionada(self):

        nro = self._get_selected_oc()

        if not nro:

            messagebox.showwarning("Selecciona una fila", "Elegí una OC para eliminar.")

            return

        if not messagebox.askyesno("Confirmar eliminación", f"¿Seguro que querés eliminar la OC {nro} y sus ítems?"):

            return

        conn = db.connect("pedidos")

        try:

            cur = conn.cursor()

            cur.execute("DELETE FROM orden_compra WHERE nro_oc = %s", (nro,))

            conn.commit()

        except Exception as e:

            messagebox.showerror("Error al eliminar", str(e))

            return

        finally:

            conn.close()

        self.lbl_info.config(text="NRO: -   |   FECHA: -   |   ITEMS: -   |   SUC: -")

        self.tree_oc_detalle.delete(*self.tree_oc_detalle.get_children())

        self.refresh_resumenes()

        self.refresh_pendientes()



    def _mostrar_texto_pedido(self):
        """
        Abre un cuadro de texto listo para copiar con destino, descripcion y cantidad.
        """
        nro = self._get_selected_oc()
        if not nro:
            messagebox.showwarning('Selecciona una fila', 'Elegi una OC para generar el texto.')
            return

        conn = db.connect("pedidos")
        try:
            cur = conn.cursor()
            cur.execute('SELECT sucursal FROM orden_compra WHERE nro_oc=%s;', (nro,))
            row_oc = cur.fetchone()
            destino = (row_oc['sucursal'] or '').strip() if row_oc else ''
            cur.execute("""
                SELECT oi.linea,
                       TRIM(CAST(oi.descripcion AS TEXT)) AS descripcion,
                       COALESCE(oi.cantidad, 0) AS cantidad
                FROM orden_item oi
                JOIN orden_compra oc ON oc.id = oi.oc_id
                WHERE oc.nro_oc = %s AND oi.descripcion IS NOT NULL
                ORDER BY oi.linea ASC, oi.descripcion ASC;
            """, (nro,))
            items = cur.fetchall()
        finally:
            conn.close()

        if not items:
            messagebox.showinfo('Sin datos', 'No hay items para esta OC.')
            return

        header = f"Destino: {destino or '-'}   |   OC: {nro}"
        col_header = f"{'Cantidad':>10} {'-':<1} {'Descripcion':<60}"
        lines = [header, '-' * max(len(header), len(col_header)), col_header]
        for r in items:
            desc_raw = (r['descripcion'] or '').strip()
            # Limpieza agresiva: sacar "el cacique", "cacique" y "el" sobrante, normalizado a minúsculas
            desc = desc_raw.lower()
            desc = desc.replace("el cacique", "").replace("cacique", "")
            desc = re.sub(r"\\bel\\b", " ", desc)
            desc = " ".join(desc.split())
            cant = int(round(r['cantidad'] or 0))
            lines.append(f"{cant:>10} - {desc:<60.60}")
        texto = "\n".join(lines)

        dlg = tk.Toplevel(self)
        dlg.title('Texto para copiar')
        dlg.geometry('900x500')
        dlg.transient(self)
        dlg.grab_set()

        txt = tk.Text(dlg, wrap='none', font=('Consolas', 10))
        txt.pack(fill='both', expand=True, padx=8, pady=8)
        txt.insert('1.0', texto)
        txt.focus_set()

        btns = ttk.Frame(dlg)
        btns.pack(fill='x', padx=8, pady=(0,8))
        def _copiar():
            try:
                self.clipboard_clear()
                self.clipboard_append(txt.get('1.0', 'end-1c'))
                messagebox.showinfo('Copiado', 'Texto copiado al portapapeles.')
            except Exception as e:
                messagebox.showerror('Error', str(e))
        ttk.Button(btns, text='Copiar al portapapeles', command=_copiar).pack(side='left')
        ttk.Button(btns, text='Cerrar', command=dlg.destroy).pack(side='left', padx=6)


    def refresh_pendientes(self):

        """

        Suma paquetes de TODAS las OCs no completadas, agrupando por descripción EXACTA,

        y para cada ítem consulta fraccionadora.db para mostrar DISPONIBLE vs NECESARIO

        con colores: verde (ok), rojo (falta), ámbar (desconocido).

        """

        # 1) Traer pendientes (necesario) desde nuestra DB

        conn = db.connect("pedidos")
        try:

            cur = conn.cursor()

            cur.execute("""

                SELECT

                    TRIM(CAST(oi.descripcion AS TEXT)) AS descripcion,

                    COALESCE(SUM(oi.cantidad),0) AS necesario

                FROM orden_item oi

                JOIN orden_compra oc ON oc.id = oi.oc_id

                WHERE oc.completada = 0

                AND oi.descripcion IS NOT NULL

                AND COALESCE(oi.enviado,0) = 0

                GROUP BY descripcion

                HAVING COALESCE(SUM(oi.cantidad),0) > 0

            """)

            rows = cur.fetchall()

            # Total Gs pendientes (suma de monto_total de OCs no completadas)
            cur.execute("""
                SELECT COALESCE(SUM(oc.monto_total),0)
                FROM orden_compra oc
                WHERE oc.completada = 0
            """)
            total_pend_gs = cur.fetchone()[0]
            try:
                total_fmt = f"{float(total_pend_gs or 0):,.0f}".replace(",", ".")
            except Exception:
                total_fmt = "0"
            if hasattr(self, "var_total_pend_gs"):
                self.var_total_pend_gs.set(total_fmt)

            cur.execute("""
                SELECT
                    COALESCE(NULLIF(TRIM(oc.sucursal), ''), 'Sin sucursal') AS sucursal,
                    COUNT(DISTINCT oc.id) AS ocs_total,
                    COALESCE(SUM(COALESCE(oc.monto_total, 0)), 0) AS monto_pendiente
                FROM orden_compra oc
                WHERE COALESCE(oc.completada,0) = 0
                GROUP BY COALESCE(NULLIF(TRIM(oc.sucursal), ''), 'Sin sucursal')
                ORDER BY COALESCE(NULLIF(TRIM(oc.sucursal), ''), 'Sin sucursal');
            """)
            rows_sucursal_base = cur.fetchall()

            cur.execute("""
                SELECT
                    COALESCE(NULLIF(TRIM(oc.sucursal), ''), 'Sin sucursal') AS sucursal,
                    TRIM(CAST(oi.descripcion AS TEXT)) AS descripcion,
                    COALESCE(oi.cantidad, 0) AS necesario
                FROM orden_item oi
                JOIN orden_compra oc ON oc.id = oi.oc_id
                WHERE COALESCE(oc.completada,0) = 0
                  AND COALESCE(oi.enviado,0) = 0
                  AND oi.descripcion IS NOT NULL
                ORDER BY COALESCE(NULLIF(TRIM(oc.sucursal), ''), 'Sin sucursal'), oi.linea, oi.id;
            """)
            rows_sucursal_items = cur.fetchall()

        finally:

            conn.close()



        # 2) Abrir fraccionadora.db una sola vez

        try:

            fr_conn = db.connect("fraccionadora")

        except Exception:

            fr_conn = None



        # 3) Pintar la tabla con colores y diagnóstico en la descripción

        self.tree_pend.delete(*self.tree_pend.get_children())



        # Ordenar similar a antes (familia + gramaje)

        import re

        def _parse_gramaje(desc: str) -> int:

            s = desc.lower()

            m = re.search(r'(\d+)\s*(?:kg|kilo|kilos)\b', s)

            if m: return int(m.group(1)) * 1000

            m = re.search(r'(\d+)\s*(?:g|gr|gramo|gramos)\b', s)

            if m: return int(m.group(1))

            m = re.search(r'\*\s*(\d{2,4})\b', s)

            if m: return int(m.group(1))

            return 10**9

        def _familia(desc: str) -> str:

            m = re.search(r'\d', desc)

            base = desc if not m else desc[:m.start()]

            return base.strip().upper()



        datos = sorted(rows, key=lambda r: (_familia(r["descripcion"]), _parse_gramaje(r["descripcion"]), r["descripcion"].upper()))
        productos_ref = _productos_referencia()
        resumen_bolsas = {nombre: 0 for nombre in productos_ref}
        resumen_bolsas.setdefault("Otros", 0)
        stock_bolsas = _stock_bolsas_por_producto(fr_conn)


        for r in datos:

            desc_base = r["descripcion"]

            desc = desc_base

            nec  = int(r["necesario"] or 0)


            tag = ("desconocido",)

            disp_txt = ""

            disp_val = None

            producto_ref = None

            if fr_conn is not None:

                estado, info, disp, prod_nombre = _mapear_y_chequear_item(desc_base, nec, fr_conn)

                tag = (estado,)

                disp_val = disp

                disp_txt = "" if disp is None else str(disp)

                producto_ref = _producto_por_desc(prod_nombre)

                # agrega el diagnóstico al texto de la descripción

                if info:

                    desc = f"{desc}  -  [{info}]"

            if not producto_ref:

                producto_ref = _producto_por_desc(desc_base)

            if producto_ref not in resumen_bolsas:

                producto_ref = "Otros"


            bolsas = _bolsas_necesarias(desc_base, nec, disp_val)

            bolsas_txt = "" if bolsas is None else str(bolsas)

            if bolsas is not None:

                resumen_bolsas[producto_ref] = resumen_bolsas.get(producto_ref, 0) + bolsas

            self.tree_pend.insert("", "end", values=(desc, nec, disp_txt, bolsas_txt), tags=tag)


        self._apply_treeview_striping(self.tree_pend)

        if hasattr(self, "var_bolsas"):

            for key, var in self.var_bolsas.items():

                var.set(str(resumen_bolsas.get(key, 0)))

        if hasattr(self, "lbl_bolsas_estado"):

            for nombre, lbl in self.lbl_bolsas_estado.items():

                nec = resumen_bolsas.get(nombre, 0)

                if stock_bolsas is None:

                    lbl.config(text="SIN DATOS", fg="#a17c00")

                    continue

                disp = stock_bolsas.get(nombre, 0.0)

                if disp >= nec:

                    lbl.config(text=f"DISP {disp:.1f} >= NEC {nec}", fg="#0a7d28")

                else:

                    lbl.config(text=f"DISP {disp:.1f} < NEC {nec}", fg="#c1121f")

        self._render_pendientes_sucursal(rows_sucursal_base, rows_sucursal_items, fr_conn)

        if fr_conn is not None:

            fr_conn.close()

    def _render_pendientes_sucursal(self, rows_sucursal_base, rows_sucursal_items, fr_conn):
        if not hasattr(self, "frm_pend_sucursales"):
            return

        for child in self.frm_pend_sucursales.winfo_children():
            child.destroy()

        if not rows_sucursal_base:
            ttk.Label(
                self.frm_pend_sucursales,
                text="Sin datos por sucursal.",
                foreground="#666666",
            ).pack(anchor="w")
            return

        detalle_por_sucursal = {}
        for row in rows_sucursal_items or []:
            sucursal = str(row["sucursal"] or "Sin sucursal")
            detalle_por_sucursal.setdefault(sucursal, []).append(row)

        total_items = 0
        total_ok = 0
        total_falta = 0
        total_desconocidos = 0
        total_ocs = 0
        total_monto = 0.0

        cols = min(3, max(1, len(rows_sucursal_base)))
        for col in range(cols):
            self.frm_pend_sucursales.columnconfigure(col, weight=1, uniform="pend_suc")

        self._pend_sucursal_vars = {}
        for idx, row in enumerate(rows_sucursal_base):
            sucursal = str(row["sucursal"] or "Sin sucursal")
            fila, col = divmod(idx, cols)
            items = detalle_por_sucursal.get(sucursal, [])
            items_total = len(items)
            items_ok = 0
            items_falta = 0
            items_desconocidos = 0

            for item in items:
                if fr_conn is None:
                    items_desconocidos += 1
                    continue
                estado, _info, _disp, _prod = _mapear_y_chequear_item(
                    item["descripcion"],
                    int(item["necesario"] or 0),
                    fr_conn,
                )
                if estado == "ok":
                    items_ok += 1
                elif estado == "falta":
                    items_falta += 1
                else:
                    items_desconocidos += 1

            pct_ok = (items_ok / items_total * 100.0) if items_total > 0 else 0.0
            pct_falta = (items_falta / items_total * 100.0) if items_total > 0 else 0.0
            pct_desconocidos = (items_desconocidos / items_total * 100.0) if items_total > 0 else 0.0

            box = ttk.LabelFrame(
                self.frm_pend_sucursales,
                text=f"{sucursal}  |  Listo: {pct_ok:.1f}%",
                padding=10,
            )
            box.grid(row=fila, column=col, sticky="nsew", padx=5, pady=5)

            ocs_total = int(row["ocs_total"] or 0)
            monto_pendiente_raw = float(row["monto_pendiente"] or 0)
            monto_pendiente = self._fmt_monto(monto_pendiente_raw)

            total_ocs += ocs_total
            total_items += items_total
            total_ok += items_ok
            total_falta += items_falta
            total_desconocidos += items_desconocidos
            total_monto += monto_pendiente_raw

            resumen = (
                f"OCs pendientes: {ocs_total}\n"
                f"Listos para enviar: {items_ok}/{items_total} ({pct_ok:.1f}%)\n"
                f"Sin stock suficiente: {items_falta}/{items_total} ({pct_falta:.1f}%)\n"
                f"Sin mapa/stock: {items_desconocidos}/{items_total} ({pct_desconocidos:.1f}%)\n"
                f"Total pendiente: {monto_pendiente}"
            )
            lbl = ttk.Label(box, text=resumen, justify="left", font=("Segoe UI", 10, "bold"))
            lbl.pack(anchor="w")

        pct_total_ok = (total_ok / total_items * 100.0) if total_items > 0 else 0.0
        pct_total_falta = (total_falta / total_items * 100.0) if total_items > 0 else 0.0
        pct_total_desconocidos = (total_desconocidos / total_items * 100.0) if total_items > 0 else 0.0
        estado_global = "SI" if total_items > 0 and total_ok == total_items else "NO"
        resumen_global = (
            f"TOTAL SUCURSALES | OCs pendientes: {total_ocs} | "
            f"Items listos: {total_ok}/{total_items} ({pct_total_ok:.1f}%) | "
            f"Con falta: {total_falta}/{total_items} ({pct_total_falta:.1f}%) | "
            f"Sin mapa/stock: {total_desconocidos}/{total_items} ({pct_total_desconocidos:.1f}%) | "
            f"Se puede enviar todo: {estado_global} | "
            f"Total pendiente: {self._fmt_monto(total_monto)}"
        )
        self.var_pend_total_general.set(resumen_global)





    def select_pdf(self):

        path = filedialog.askopenfilename(

            title="Selecciona PDF de Orden de Compra",

            filetypes=[("PDF", "*.pdf")]

        )

        if path:

            self.var_pdf.set(path)



    def parse_and_save(self):

        pdf_path = Path(self.var_pdf.get())

        if not pdf_path.exists():

            messagebox.showerror("Error", "Selecciona un PDF válido.")

            return

        try:

            result = parse_pdf(pdf_path)

            meta, items = result["meta"], result["items"]



            self.var_oc.set(meta.get("nro_oc", ""))

            self.var_suc.set(meta.get("sucursal", ""))

            self.var_fecha.set(meta.get("fecha_pedido", ""))



            # Carga de items con cantidad ajustada

            # Carga de items con cantidad ajustada para la UI

            self.items_full = []

            for i, it in enumerate(result.get("items", []), start=1):

                cant_adj = ajustar_cantidad(it.get("descripcion", ""), it.get("cantidad"))

                self.items_full.append({"linea": i, "desc": it.get("descripcion"), "qty": cant_adj})



            # pintar tabla

            for rid in self.tree.get_children():

                self.tree.delete(rid)

            for r in self.items_full:

                self.tree.insert("", "end", values=(r["linea"], r["desc"], r["qty"]))

            self._apply_treeview_striping(self.tree)



            # GUARDAR EN DB: pasa los items crudos, insert_oc ya AJUSTA internamente

            oc_id = insert_oc(meta, items)



            # ==== Generar XLSX automáticamente, sin diálogos ====

            if not TEMPLATE_XLSX.exists():

                messagebox.showerror("Plantilla faltante",

                                    f"No encontré la plantilla:\n{TEMPLATE_XLSX}")

                return



            OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

            oc_num = (meta.get("nro_oc") or "SIN_OC").strip()

            timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")

            out_path = OUTPUT_DIR / f"NOTA_REMISION_{oc_num}_{timestamp}.xlsx"



            # Construir items para la plantilla con lo que ve la UI (ya ajustado e integer)

            items_tpl = [{"descripcion": r["desc"], "cantidad": r["qty"]} for r in self.items_full]

            # Peso total del pedido (según paquetes ya ajustados)

            peso_kg, _detalle = calcular_peso_total_items(items_tpl)



            try:

                escritos = cargar_cantidades_en_plantilla(TEMPLATE_XLSX, out_path, items_tpl)

            except Exception as e:

                messagebox.showerror("Error al generar XLSX", str(e))

                return



            messagebox.showinfo(

                "Listo",

                f"Guardado en SQLite (oc_id={oc_id}).\n"

                f"Plantilla generada: {out_path}\n"

                f"Filas actualizadas: {escritos}\n"

                f"Peso total del pedido: {peso_kg} kg"



            )



            messagebox.showinfo("Listo", f"Guardado en SQLite (oc_id={oc_id}).")

        except Exception as e:

            messagebox.showerror("Falló el parseo", str(e))

        self.refresh_resumenes()

        self.refresh_pendientes()



if __name__ == "__main__":

    App().mainloop()
